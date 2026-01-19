import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_file = 'BuiltInCategoriesRevit2024.xlsx'
output_xlsx = 'Revit2024_VisibilityGraphics_ModelCategories.xlsx'
output_csv = 'Revit2024_VisibilityGraphics_ModelCategories.csv'
output_md = 'Revit2024_VisibilityGraphics_ModelCategories.md'

# Load data
print(f"Reading {input_file}...")
df = pd.read_excel(input_file)

# Filter
# Drop NaN names
df = df.dropna(subset=['English_USA (ENU)'])

# Filter out <...>
df = df[~df['English_USA (ENU)'].str.startswith('<')]

# Filter out specific internal categories
df = df[df['English_USA (ENU)'] != "Internal Object Styles"]

# Filter out internal load/structural categories that appear internal
internal_load_keywords = [
    'Internal Area Load', 'Internal Line Load', 'Internal Point Load',
    'Internal Area Loads', 'Internal Line Loads', 'Internal Point Loads',
    'Structural Internal Loads'
]
for keyword in internal_load_keywords:
    df = df[~df['English_USA (ENU)'].str.contains(keyword)]

# Reset index
df = df.reset_index(drop=True)

# Prepare output DataFrame
output_df = pd.DataFrame()
output_df['Visibility'] = ['TRUE'] * len(df)
output_df['Category'] = df['English_USA (ENU)']
output_df['Projection/Surface - Lines'] = [''] * len(df)
output_df['Projection/Surface - Patterns'] = [''] * len(df)
output_df['Transparency'] = [0] * len(df)
output_df['Cut - Lines'] = [''] * len(df)
output_df['Cut - Patterns'] = [''] * len(df)
output_df['Halftone'] = ['FALSE'] * len(df)
output_df['Detail Level'] = ['By View'] * len(df)
output_df['ID'] = df['ID']

# Sort by Category
output_df = output_df.sort_values(by='Category')

# 1. Generate CSV
print(f"Generating {output_csv}...")
output_df.to_csv(output_csv, index=False)

# 2. Generate Markdown
print(f"Generating {output_md}...")
# We use tabulate under the hood
markdown_content = output_df.to_markdown(index=False, tablefmt="github")
with open(output_md, "w") as f:
    f.write(markdown_content)

# 3. Generate Excel
print(f"Generating {output_xlsx}...")
# Use pandas to write basic Excel first
output_df.to_excel(output_xlsx, index=False, sheet_name='Model Categories')

# Apply formatting using openpyxl
wb = load_workbook(output_xlsx)
ws = wb.active

# Apply column widths and hiding
# Columns are 1-indexed in openpyxl
# A: Visibility
# B: Category
# C: Projection/Surface - Lines
# D: Projection/Surface - Patterns
# E: Transparency
# F: Cut - Lines
# G: Cut - Patterns
# H: Halftone
# I: Detail Level
# J: ID

ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 20
ws.column_dimensions['G'].width = 20
ws.column_dimensions['H'].width = 10
ws.column_dimensions['I'].width = 15

# Hide ID column (Column J is the 10th column)
ws.column_dimensions['J'].hidden = True

wb.save(output_xlsx)
print(f"Done. Generated 3 files with {len(output_df)} rows each.")
